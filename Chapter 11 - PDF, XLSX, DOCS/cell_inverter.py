import openpyxl, sys

wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.active

sheet_copy = []
for r in range(1, sheet.max_row + 1):
    row_copy = []
    for c in range(1, sheet.max_column + 1):
        row_copy.append(sheet.cell(row=r, column=c).value)
    sheet_copy.append(row_copy)

wb2 = openpyxl.Workbook()
sheet2 = wb2.active
for r in range(len(sheet_copy)):
    for c in range(len(sheet_copy[0])):
        sheet2.cell(row=c+1, column=r+1).value = sheet_copy[r][c]

wb2.save('inverted_copy.xlsx')
