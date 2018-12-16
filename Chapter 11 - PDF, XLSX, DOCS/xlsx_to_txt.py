import openpyxl, os, sys

file = sys.argv[1]

excel_file = openpyxl.load_workbook(file)
sheet = excel_file.active

for c in range(1, sheet.max_column + 1):
    txt_file = open('file {}.txt'.format(c), 'w')
    for r in range(1, sheet.max_row + 1):
        value = sheet.cell(row=r, column=c).value
        if value:
            txt_file.write('{}\n'.format(value))
        else:
            pass
    txt_file.close()


