import openpyxl, sys

if len(sys.argv) == 4:
    rou = sys.argv[1]
    num = sys.argv[2]
    file = sys.argv[3]

    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    sheet_part = []
    for r in range(int(rou), sheet.max_row + 1):
        row_to_copy = []
        for c in range(1, sheet.max_column + 1):
            row_to_copy.append(sheet.cell(row=r, column=c).value)
        sheet_part.append(row_to_copy)
    for r in range(int(rou), int(rou) + int(num)):
        for c in range(1, sheet.max_column + 1):
            sheet.cell(row=r, column=c).value = ''
    for r in range(int(rou) + int(num), sheet.max_row + int(num) + 1):
        for c in range(1, sheet.max_column + 1):
            sheet.cell(row=r, column=c).value = sheet_part[r - int(rou) - int(num)][c - 1]

    wb.save('mul_copy.xlsx')
