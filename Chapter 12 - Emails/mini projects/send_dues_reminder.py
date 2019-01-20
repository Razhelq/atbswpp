# send_dues_reminders.py - Sends emails based on payment status in spreadsheet.


import openpyxl, smtplib, sys


# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('..\\..\\automate_online-materials\\duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

last_col = sheet.max_column
lastest_month = sheet.cell(row=1, column=last_col).value

unpaid_members = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

smtp_obj = smtplib.SMTP()
smtp_obj.connect('smtp.gmail.com', '587')
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('email_address', 'password')

for name, email in unpaid_members.items():
    body = "Subject: {} dues unpaid. \nDear {},\n Records show that you have not paid dues for {}. " \
           "Please make this payment as soon as possible. Thank you!".format(lastest_month, name, lastest_month)
    print('Sending email to {}...'.format(name))
    print(email)
    print(body)
    send_mail_status = smtp_obj.sendmail('email_address', email, body)

    if send_mail_status != {}:
        print('There was a problem sending email to {}: {}'.format(email, send_mail_status))

smtp_obj.quit()