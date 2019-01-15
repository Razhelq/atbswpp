import openpyxl, sys
from openpyxl.styles import Font


num = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1, num + 1):
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=1, column=i + 1).font = Font(bold=True)

    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=i + 1, column=1).font = Font(bold=True)

cs = 1
for c in range(2, num + 2):
    for r in range(2, num + 2):
        sheet.cell(row=r, column=c).value = int(sheet.cell(row=1, column=c).value) * int(sheet.cell(row=r, column=1).value)
        cs = + 1

wb.save('multiplication_table.xlsx')